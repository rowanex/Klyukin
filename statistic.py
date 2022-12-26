import openpyxl
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter
import matplotlib
import matplotlib.pyplot as plt
import numpy as np

import csv
import math
import os
import re

from matplotlib.axes import Axes
from jinja2 import Template
import pdfkit

matplotlib.use('TkAgg')
plt.rcParams.update({'font.size': 8})

dic_naming = {
    'name': 'Название',
    'description': 'Описание',
    'key_skills': 'Навыки',
    'experience_id': 'Опыт работы',
    'premium': 'Премиум-вакансия',
    'employer_name': 'Компания',
    'salary': 'Оклад',
    'area_name': 'Название региона',
    'published_at': 'Дата публикации вакансии',
}

dic_naming_filter_helped = {
    'Название': 0,
    'Описание': 1,
    'Навыки': 2,
    'Опыт работы': 3,
    'Премиум-вакансия': 4,
    'Компания': 5,
    'Оклад': '6, 7, 8',
    'Идентификатор валюты оклада': 9,
    'Название региона': 10,
    'Дата публикации вакансии': 11
}

dic_naming_filter_helped_reversed = {
    0: 'Название',
    1: 'Описание',
    2: 'Навыки',
    3: 'Опыт работы',
    4: 'Премиум-вакания',
    5: 'Компания',
    '6, 7, 8': 'Оклад',
    9: 'Идентификатор валюты оклада',
    10: 'Название региона',
    11: 'Дата публикации вакансии'
}

dic_experience = {
    "noExperience": "Нет опыта",
    "between1And3": "От 1 года до 3 лет",
    "between3And6": "От 3 до 6 лет",
    "moreThan6": "Более 6 лет"
}

dic_currency = {
    "AZN": "Манаты",
    "BYR": "Белорусские рубли",
    "EUR": "Евро",
    "GEL": "Грузинский лари",
    "KGS": "Киргизский сом",
    "KZT": "Тенге",
    "RUR": "Рубли",
    "UAH": "Гривны",
    "USD": "Доллары",
    "UZS": "Узбекский сум",
}
currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}
dic_true_false = {
    'False': 'Нет',
    'True': 'Да',
    'FALSE': 'Нет',
    'TRUE': 'Да'
}
dic_salary_gross = {
    'Нет': 'С вычетом налогов',
    'Да': 'Без вычета налогов'
}


class Vacancy:
    """Класс для представления вакансии

       Attributes:
           name (str): Название вакансии
           salary (int): Средняя зарплата
           area_name (str): Название региона
           published_at (str): Дата публикации вакансии
       """
    def __init__(self, name, salary, area_name, published_at):
        """Инициализирует объект Vacancy

        Args:
            name (str): Название вакансии
            salary (str or int or float): Средняя зарплата
            area_name (str): Название региона
            published_at (str): Дата публикации вакансии
        """
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at


class Salary:
    """ Класс для представления зарплаты

    Attributes:
        salary_from (int): Нижняя граница вилки оклада
        salary_to (int): Верхняя граница вилки оклада
        salary_currency (str): Идентификатор валюты оклада
    """
    def __init__(self, salary_from, salary_to, salary_currency):
        """ Инициализирует объект Salary.

        Args:
            salary_from (str or int or float): Нижняя граница вилки оклада
            salary_to (str or int or float): Верхняя граница вилки оклада
            salary_currency (str): Идентификатор валюты оклада
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency


class Report:
    """ Класс для создания репорта по значениям в соответствии с ТЗ

    Attributes:
        salary_by_years_dict(dict{int: int}): Средняя заплата вакансий по годам
        number_of_vacancies_by_year_dict(dict{int: int}): Количество вакансий по годам
        salary_for_vacancy_by_years_dict(dict{int: int}): Зарпалата по годам для выбранной профессии
        number_of_vacancies_by_year_for_vac_dict(dict{int: int}): Колличество вакансий по годам выбранной пофесиий
        salary_by_city_dict(dict{str: int}): Средняя зарплата по городам
        number_of_vacancies_by_city_dict(dict{str: int or str}): Количество вакансий по городам
        number_of_vacancies_by_city_dict_percent(dict{str: str}): Количество вакансий по городам по процентам
        profession_name(str): Название выбранной профессии
        sheet_1_headers (list(str)): Названия колонок для формирования первой таблицы (статистика по годам)
        sheet_2_headers (list(str)): Названия колонок для формирования второй таблицы (статистика по городам)
        sheet_years_rows (list[list[str]]): Значения для первой таблицы (статистика по годам)
        sheet_cities_rows (list[list[str]]): Значения для второй таблицы (статистика по городам)
    """
    def __init__(self, salary_by_years_dict, number_of_vacancies_by_year_dict, salary_for_vacancy_by_years_dict,
                 number_of_vacancies_by_year_for_vac_dict, salary_by_city_dict,
                 number_of_vacancies_by_city_dict, number_of_vacancies_by_city_dict_percent, profession_name):
        """ Инициализация данных для статистики вакансий и выбранной профессии

        Args:
            salary_by_years_dict(dict{int: int}): Средняя заплата вакансий по годам
            number_of_vacancies_by_year_dict(dict{int: int}): Количество вакансий по годам
            salary_for_vacancy_by_years_dict(dict{int: int}): Зарпалата по годам для выбранной профессии
            number_of_vacancies_by_year_for_vac_dict(dict{int: int}): Колличество вакансий по годам выбранной пофесиий
            salary_by_city_dict(dict{str: int}): Средняя зарплата по городам
            number_of_vacancies_by_city_dict(dict{str: int or str}): Количество вакансий по городам
            number_of_vacancies_by_city_dict_percent(dict{str: str}): Количество вакансий по городам по процентам
            profession_name(str): Название выбранной профессии
        """
        self.salary_by_years_dict = salary_by_years_dict
        self.number_of_vacancies_by_year_dict = number_of_vacancies_by_year_dict
        self.salary_for_vacancy_by_years_dict = salary_for_vacancy_by_years_dict
        self.number_of_vacancies_by_year_for_vac_dict = number_of_vacancies_by_year_for_vac_dict
        self.salary_by_city_dict = salary_by_city_dict
        self.number_of_vacancies_by_city_dict = number_of_vacancies_by_city_dict
        self.number_of_vacancies_by_city_dict_percent = number_of_vacancies_by_city_dict_percent
        self.profession_name = profession_name
        self.sheet_1_headers = ["Год", "Средняя зарплата", f"Средняя зарплата -{self.profession_name}",
                                "Количество вакансий", f"Количество вакансий -{self.profession_name}"]
        self.sheet_2_headers = ["Город", "Уровень зарплат", " ", "Город", "Доля вакансий"]
        self.sheet_years_rows = self.get_table_rows([list(self.salary_by_years_dict.keys()), list(self.salary_by_years_dict.values()),
                               list(self.salary_for_vacancy_by_years_dict.values()), list(self.number_of_vacancies_by_year_dict.values()),
                               list(self.number_of_vacancies_by_year_for_vac_dict.values())])
        self.sheet_cities_rows = self.get_table_rows([list(self.salary_by_city_dict.keys()), list(self.salary_by_city_dict.values()),
                           ["" for _ in self.salary_by_city_dict.keys()], list(self.number_of_vacancies_by_city_dict.keys()),
                           list(self.number_of_vacancies_by_city_dict_percent.values())])

    @staticmethod
    def fill_column(sheet, data, letters):
        """Функция заполнения колонок таблицы статистики данными

        Attributes:
            sheet(Workbook): Лист таблицы
            data(list[{items}]): Данные для записи
            letters(str): Колонки таблицы
        """
        for i, item in enumerate(data):
            for row, (key, value) in enumerate(item.items(), start=2):
                sheet[f'{letters[i]}{row}'] = value

    def fill_year_sheet(self, sheet_statistic_by_year):
        """Функция заполнения листа статистикой по годам

        Attributes:
            sheet_statistic_by_year(Workbook): лист книги xlsx для статистики по годам
        Returns:
            Workbook: лист книги xlsx для статистики по годам
        """
        sheet_statistic_by_year.title = 'Статистика по годам'
        sheet_statistic_by_year['A1'] = 'Год'
        sheet_statistic_by_year['B1'] = 'Средняя зарплата'
        sheet_statistic_by_year['C1'] = f'Средняя зарплата - {self.profession_name}'
        sheet_statistic_by_year['D1'] = 'Количество вакансий'
        sheet_statistic_by_year['E1'] = f'Количество вакансий - {self.profession_name}'
        column_letter_year = 'ABCDE'
        for letter in column_letter_year:
            sheet_statistic_by_year[f'{letter}1'].font = Font(bold=True)
        for row, (key, value) in enumerate(self.salary_by_years_dict.items(), start=2):
            sheet_statistic_by_year[f'A{row}'] = key
        year = [self.salary_by_years_dict, self.salary_for_vacancy_by_years_dict, self.number_of_vacancies_by_year_dict, self.number_of_vacancies_by_year_for_vac_dict]
        Report.fill_column(sheet_statistic_by_year, year, column_letter_year[1:])
        return sheet_statistic_by_year

    def fill_city_sheet(self, sheet_statistic_by_city):
        """Функция заполнения листа статистикой по годам

        Attributes:
            sheet_statistic_by_city(Workbook): лист книги xlsx для статистики по городам
        Returns:
            Workbook: лист книги xlsx для статистики по городам
        """
        sheet_statistic_by_city['A1'] = 'Город'
        sheet_statistic_by_city['B1'] = 'Уровень зарплат'
        sheet_statistic_by_city['D1'] = 'Город'
        sheet_statistic_by_city['E1'] = 'Доля вакансий'
        column_letter_city = 'ABDE'
        for letter in column_letter_city:
            sheet_statistic_by_city[f'{letter}1'].font = Font(bold=True)
        for row, (key, value) in enumerate(self.salary_by_city_dict.items(), start=2):
            sheet_statistic_by_city[f'A{row}'] = key
            sheet_statistic_by_city[f'D{row}'] = key
            sheet_statistic_by_city[f'B{row}'] = value
        for row, (key, value) in enumerate(self.number_of_vacancies_by_city_dict_percent.items(), start=2):
            sheet_statistic_by_city[f'E{row}'] = value
        return sheet_statistic_by_city

    @staticmethod
    def get_border(sheet, border):
        """Функция создания линий гранц таблицы

        Attributes:
            sheet(Workbook): Созданный лист таблицы
            border(Side): Граница
        """
        for row in sheet.columns:
            for cell in row:
                if cell.value:
                    cell.border = Border(top=border, left=border, right=border, bottom=border)
        return sheet

    @staticmethod
    def set_width_cells(sheet):
        """Функция автоматической регулировки ширины коллонки

        Attributes:
            sheet(Workbook): Лист таблицы
        """
        column_widths = []
        for row in sheet:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell.value)) > column_widths[i]:
                        column_widths[i] = len(str(cell.value)) + 1
                else:
                    column_widths += [len(str(cell.value)) + 1]
        for i, column_width in enumerate(column_widths, 1):
                sheet.column_dimensions[get_column_letter(i)].width = (column_width)
                if get_column_letter(i) == 'C' and sheet.title == 'Статистика по городам':
                    sheet.column_dimensions[get_column_letter(i)].width = 1.5
                else:
                    sheet.column_dimensions[get_column_letter(i)].width = (column_width + 2)

    def generate_excel(self):
        """Функция генерирует таблицу в формате xlxs

        Returns:
            book(Workbook): Созданная таблица, чтобы потом сохранить её в нужном нам имени
        """
        book = openpyxl.Workbook()
        sheet_statistic_by_year = book.active
        sheet_statistic_by_city = book.create_sheet('Статистика по городам')
        sheet_statistic_by_year = Report.fill_year_sheet(self, sheet_statistic_by_year)
        sheet_statistic_by_city = Report.fill_city_sheet(self, sheet_statistic_by_city)
        Report.get_border(sheet_statistic_by_year, Side(border_style='thin', color='000000'))
        Report.get_border(sheet_statistic_by_city, Side(border_style='thin', color='000000'))
        Report.set_width_cells(sheet_statistic_by_year)
        Report.set_width_cells(sheet_statistic_by_city)
        return book

    def draw_horizontal_diagram(self, ax: Axes, keys_general, values_general, values_vacancy, label_general, label_vacancy, title):
        """ Функция отрисовывает горизонтальные диаграммы

        Attributes:
            ax: Axes(Matplotlib.axes): Экземпляр класса осей
            keys_general(list[str]): значения для тиков и меток тиков
            values_general(list[str]): значения для основной статистики (столбца для всех вакансий)
            values_general(list[str]): значения для статистики по выбранной вакансии(столбца для выбранной вакансии)
            label_general(str): значения для обозначения основной статистики (лэйбл основных слобцов)
            label_vacancy(str): значения для обозначения статистики по вакансии (лэйбл слобцов для выбранной вакансии)
            title(str): заголовок для диаграммы
        """
        x = np.arange(len(keys_general))
        ax.bar(x - 0.25, list(values_general), width=0.5, label=label_general)
        ax.bar(x + 0.25, list(values_vacancy), width=0.5, label=label_vacancy)
        ax.set_xticks(range(len(self.salary_by_years_dict)), list(keys_general), rotation=90)
        ax.legend()
        ax.set_title(title)
        ax.grid(visible=True, axis="y")

    def draw_vertical_diagram(self, ax: Axes, title):
        """ Функция отрисовывает вертикальные диаграммы

        Attributes:
            ax: Axes(Matplotlib.axes): Экземпляр класса осей
            title(str): заголовок для диаграммы
        """
        keys = [key.replace(" ", "\n").replace("-", "-\n") for key in reversed(list(self.salary_by_city_dict.keys()))]
        values = [x for x in reversed(list(self.salary_by_city_dict.values()))]
        ax.barh(keys, values)
        ax.grid(visible=True, axis='x')
        ax.set_title(title)

    def draw_pie_diogramm(self, ax: Axes, title):
        """ Функция отрисовывает круговые диаграммы

        Attributes:
            ax: Axes(Matplotlib.axes): Экземпляр класса осей
            title(str): заголовок для диаграммы
        """
        plt.rc('xtick', labelsize=6)
        dict = self.number_of_vacancies_by_city_dict
        dict["Другие"] = 1 - sum([val for val in dict.values()])
        keys = list(dict.keys())
        ax.pie(list(dict.values()), labels=keys)
        ax.axis('equal')
        ax.set_title(title)

    def generate_image(self, file_name):
        """ Функция создет и сохраняет графики по ТЗ

        Attributes:
            file_name(str): Название файла с графиками
        """
        fig, ax = plt.subplots(nrows=2, ncols=2)
        self.draw_horizontal_diagram(ax[0, 0], self.salary_by_years_dict.keys(), self.salary_by_years_dict.values(), self.salary_for_vacancy_by_years_dict.values(), "Средняя з/п", f"з/п {self.profession_name}", "Уровень зарплат по годам")
        self.draw_horizontal_diagram(ax[0, 1], self.number_of_vacancies_by_year_dict.keys(), self.number_of_vacancies_by_year_dict.values(), self.number_of_vacancies_by_year_for_vac_dict.values(), "Количество вакансий", f"Количество вакансий {self.profession_name} ", "Количество вакансий по годам")
        self.draw_vertical_diagram(ax[1, 0], "Уровень зарплат по городам")
        self.draw_pie_diogramm(ax[1, 1], "Доля вакансий по городам")
        fig.tight_layout()
        fig.savefig(file_name)

    def get_table_rows(self, columns: list):
        """ Геттер, возвращающий list колонок
        Attributes:
            columns (list) : Значение, у которого надо выбрать строки
        Returns:
            str: Переданное value в процентах
        """
        rows_list = [["" for _ in range(len(columns))] for _ in range(len(columns[0]))]
        for col in range(len(columns)):
            for cell in range(len(columns[col])):
                rows_list[cell][col] = columns[col][cell]
        return rows_list

    def generate_pdf(self, png_name):
        """Функция генерирует pdf по шаблону html_template

        Attributes:
            png_name(str): Имя файла картинки графиков
        """
        html = open("html_template.html").read()
        template = Template(html)
        keys_to_values = {
            "title": "Аналитика по зарплатам и городам для профессии " + self.profession_name,
            "png_name": png_name,
            "years_title": "Статистика по годам",
            "years_headers": self.sheet_1_headers,
            "years_rows": self.sheet_years_rows,
            "cities_title": "Статистика по городам",
            "cities_headers": self.sheet_2_headers,
            "count_columns": len(self.sheet_2_headers),
            "cities_rows": self.sheet_cities_rows
        }
        pdf_template = template.render(keys_to_values)
        config = pdfkit.configuration(wkhtmltopdf=r"C:/Program Files/wkhtmltopdf/bin/wkhtmltopdf.exe")
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": True})



class InputConect:
    """Класс для получения всей нужной информации для создания графиков и таблиц"""
    @staticmethod
    def check_file(file_name):
        """Функция проверяет файл на наличие информации

        Attributes:
            file_name(str): Имя файла с информацией о вакансифях
        """
        if os.stat(file_name).st_size == 0:
            print('Пустой файл')
            exit()
        else:
            return file_name

    @staticmethod
    def get_input_parameters():
        """Функция получает имя файла и название профессии для анализа

         :Returns
            file_name(str): название файла с информацией
            profession_name(str): имя профессии для анализа
        """
        print('Введите название файла:', end=' ')
        file_name = InputConect().check_file(input())
        print('Введите название профессии:', end=' ')
        profession_name = input()
        return file_name, profession_name

    @staticmethod
    def currency_transfer(averageSalary, currency):
        """Функция возвращает переведенную в рубли сумму
        Attributes:
            averageSalary(float): название файла с информацией
            currency(str): Идентификатор валюты оклада
        Returns:
            averageSalary * currency_to_rub[currency](float): переведенное значение суммы в рубли
        """
        return averageSalary * currency_to_rub[currency]

    @staticmethod
    def get_average_salary(vacancy):
        """Функция считает среднюю зарплату

        Attributes:
            vacancy(object): Имя файла картинки графиков
        Returns:
            currency_transfer(): возвращаем результат функции перевода в рубли
        """
        salary_from = float(vacancy.salary.salary_from)
        salary_to = vacancy.salary.salary_to
        if type(salary_to) == list:
            salary_to = float(salary_to[0])
        else:
            salary_to = float(salary_to)
        currency = vacancy.salary.salary_currency
        averageSalary = (salary_from + salary_to) / 2
        return InputConect.currency_transfer(averageSalary, currency)

    @staticmethod
    def get_years_dict_and_city_dict(vacancies_objects):
        """Функция генерирует pdf по шаблону html_template

        Attributes:
            vacancies_objects(list[object]): Лист объектов вакансий
        Returns:
            years_dict(dict{}): словарь, где ключи - года
            city_dict(dict{}): словарь, где ключи - города
        """
        years_dict = {}
        city_dict = {}
        for vacancy in vacancies_objects:
            if vacancy.published_at[0:4] not in years_dict.keys():
                years_dict[int(vacancy.published_at[0:4])] = 0
            if vacancy.area_name not in city_dict.keys():
                city_dict[vacancy.area_name] = 0
            else:
                continue
        return years_dict, city_dict

    @staticmethod
    def get_all_statistic_by_years(year, vacancies_objects, profession_name):
        """Функция получает статистику по годам

        Attributes:
            year(str): год
            vacancies_objects(list[object]): Лист объектов вакансий
            profession_name(str): Название выбранной профессии
        Returns:
            sum_salary_of_vacancies_by_year(float): зп по всем вакансиям за год
            count_of_vacancies(int): количество вакансий по всем вакансиям за год
            sum_salary_for_one_vacancy(float): зп для одной профессии за год
            count_of_vacancies_for_one_vac(int): количество вакансий по одной профессии за год
        """
        salary_of_vacancies = []
        salary_for_one_vacancy = []
        for vacancy in vacancies_objects:
            if vacancy.published_at[0:4] == str(year):
                salary_of_vacancies.append(InputConect.get_average_salary(vacancy))
            if vacancy.published_at[0:4] == str(year) and profession_name in vacancy.name:
                salary_for_one_vacancy.append(InputConect.get_average_salary(vacancy))
        count_of_vacancies = len(salary_of_vacancies)
        sum_salary_of_vacancies_by_year = math.floor(sum(salary_of_vacancies) / count_of_vacancies)
        count_of_vacancies_for_one_vac = len(salary_for_one_vacancy)
        if count_of_vacancies_for_one_vac != 0:
            sum_salary_for_one_vacancy = int(sum(salary_for_one_vacancy) / count_of_vacancies_for_one_vac)
        else:
            sum_salary_for_one_vacancy = 0
        return sum_salary_of_vacancies_by_year, count_of_vacancies, sum_salary_for_one_vacancy, count_of_vacancies_for_one_vac

    @staticmethod
    def get_all_statistic_by_city(city, vacancies_objects):
        """Функция получает статистику по городам

        Attributes:
            city(str): город
            vacancies_objects(list[object]): Лист объектов вакансий
        Returns:
            (int): зп по городам
            (int): доля вакансий по городам
        """
        salary = []
        for vacancy in vacancies_objects:
            if vacancy.area_name == city:
                salary.append(InputConect.get_average_salary(vacancy))
        if len(vacancies_objects) / 100 <= len(salary):
            return int((sum(salary)) / len(salary)), len(salary) / len(vacancies_objects)
        else:
            return 0, 0

    @staticmethod
    def print_data_for_graph(salary_by_years_dict, number_of_vacancies_by_year_dict, salary_for_vacancy_by_years_dict,
                             number_of_vacancies_by_year_for_vac_dict, sorted_salary_by_city_dict_without_zero,
                             sorted_number_of_vacancies_by_city_dict_without_zero):
        """Функция печатает статистику в соответствии с ТЗ

        Attributes:
            salary_by_years_dict(dict{}): Словарь зп по годам для всех вакансий
            number_of_vacancies_by_year_dict(dict{}): Словарь кол-ва вакансий для всех вакансий
            salary_for_vacancy_by_years_dict(dict{}): Словарь зп для выбранной профессии
            number_of_vacancies_by_year_for_vac_dict(dict{}): Словарь кол-ва вакансий для выбранной профессии
            sorted_salary_by_city_dict_without_zero(dict{}): Отсортированный словарь зп по городам без 0 значений
            sorted_number_of_vacancies_by_city_dict_without_zero(dict{}): Отсортированный словарь кол-ва вакансий по городам без 0 значений
        """
        print('Динамика уровня зарплат по годам:', salary_by_years_dict)
        print('Динамика количества вакансий по годам:', number_of_vacancies_by_year_dict)
        print('Динамика уровня зарплат по годам для выбранной профессии:', salary_for_vacancy_by_years_dict)
        print('Динамика количества вакансий по годам для выбранной профессии:',
              number_of_vacancies_by_year_for_vac_dict)
        print('Уровень зарплат по городам (в порядке убывания):', sorted_salary_by_city_dict_without_zero)
        print('Доля вакансий по городам (в порядке убывания):', sorted_number_of_vacancies_by_city_dict_without_zero)

    @staticmethod
    def get_data_for_graph(vacancies_objects, profession_name):
        """Функция получает данные для графиков

        Attributes:
            vacancies_objects(list[object]): лист объектов вакансий
            profession_name(str): название выбранной профессии
        Returns:
            salary_by_years_dict(dict{}): Словарь зп по годам для всех вакансий
            number_of_vacancies_by_year_dict(dict{}): Словарь кол-ва вакансий для всех вакансий
            salary_for_vacancy_by_years_dict(dict{}): Словарь зп для выбранной профессии
            number_of_vacancies_by_year_for_vac_dict(dict{}): Словарь кол-ва вакансий для выбранной профессии
            sorted_salary_by_city_dict_without_zero(dict{}): Сортированный словарь зп по городам без 0 значений
            sorted_number_of_vacancies_by_city_dict_without_zero(dict{}): Сортированный словарь кол-ва вакансий по городам без 0 значений
            number_of_vacancies_by_city_dict_percent(dict{}): Словарь количество вакансий по городам в процентах
        """
        years_dict, city_dict = InputConect.get_years_dict_and_city_dict(vacancies_objects)
        salary_by_years_dict = {}
        number_of_vacancies_by_year_dict = {}
        salary_for_vacancy_by_years_dict = {}
        number_of_vacancies_by_year_for_vac_dict = {}
        salary_by_city_dict = {}
        number_of_vacancies_by_city_dict = {}
        for year in years_dict.keys():
            sum_salary_of_vacancies_by_year, count_of_vacancies, \
            sum_salary_for_one_vacancy, count_of_vacancies_for_one_vac = InputConect.get_all_statistic_by_years(year,
                                                                                                                vacancies_objects,
                                                                                                                profession_name)
            salary_by_years_dict[year] = sum_salary_of_vacancies_by_year
            number_of_vacancies_by_year_dict[year] = count_of_vacancies
            salary_for_vacancy_by_years_dict[year] = sum_salary_for_one_vacancy
            number_of_vacancies_by_year_for_vac_dict[year] = count_of_vacancies_for_one_vac
        for city in city_dict.keys():
            salary_levels_by_city, share_of_vacancies_by_city = InputConect.get_all_statistic_by_city(city,
                                                                                                      vacancies_objects)
            salary_by_city_dict[city] = salary_levels_by_city
            number_of_vacancies_by_city_dict[city] = share_of_vacancies_by_city
        sorted_salary_by_city_dict = dict(
            sorted(salary_by_city_dict.items(), key=lambda item: item[1], reverse=True)[:10])
        sorted_number_of_vacancies_by_city_dict = dict(
            sorted(number_of_vacancies_by_city_dict.items(), key=lambda item: item[1], reverse=True)[:10])
        for key in sorted_number_of_vacancies_by_city_dict.keys():
            sorted_number_of_vacancies_by_city_dict[key] = round(sorted_number_of_vacancies_by_city_dict[key], 4)
        sorted_salary_by_city_dict_without_zero = {k: v for k, v in sorted_salary_by_city_dict.items() if v != 0}
        sorted_number_of_vacancies_by_city_dict_without_zero = {k: v for k, v in
                                                                sorted_number_of_vacancies_by_city_dict.items() if
                                                                v != 0}
        InputConect.print_data_for_graph(salary_by_years_dict, number_of_vacancies_by_year_dict,
                                         salary_for_vacancy_by_years_dict,
                                         number_of_vacancies_by_year_for_vac_dict,
                                         sorted_salary_by_city_dict_without_zero,
                                         sorted_number_of_vacancies_by_city_dict_without_zero)
        number_of_vacancies_by_city_dict_percent = {}
        for key in sorted_number_of_vacancies_by_city_dict.keys():
            number_of_vacancies_by_city_dict_percent[
                key] = f'{round(sorted_number_of_vacancies_by_city_dict[key] * 100, 2)}%'
        return (salary_by_years_dict, number_of_vacancies_by_year_dict, salary_for_vacancy_by_years_dict,
                number_of_vacancies_by_year_for_vac_dict, sorted_salary_by_city_dict_without_zero,
                sorted_number_of_vacancies_by_city_dict_without_zero, number_of_vacancies_by_city_dict_percent)

    @staticmethod
    def start():
        """Функция запускающая получение статистики и сохранения его в файлы по ТЗ
        """
        file_name, profession_name = InputConect().get_input_parameters()
        data_set = DataSet(file_name)
        salary_by_years_dict, number_of_vacancies_by_year_dict, salary_for_vacancy_by_years_dict,\
        number_of_vacancies_by_year_for_vac_dict, salary_by_city_dict, number_of_vacancies_by_city_dict,\
        number_of_vacancies_by_city_dict_percent = InputConect.get_data_for_graph(
            data_set.vacancies_objects, profession_name)
        statistic = Report(salary_by_years_dict, number_of_vacancies_by_year_dict, salary_for_vacancy_by_years_dict,
                       number_of_vacancies_by_year_for_vac_dict, salary_by_city_dict, number_of_vacancies_by_city_dict, number_of_vacancies_by_city_dict_percent,
                       profession_name)
        statistic.generate_excel().save('report.xlsx')
        statistic.generate_image('graph.png')
        statistic.generate_pdf('graph.png')


class DataSet:
    """ Класс для создания датасета с полученными вакансиями из csv файла

    Attributes:
        vacancies (list[list[str]]): Лист объектов вакансий
    """
    def __init__(self, file_name):
        """ Инициализирует объект DataSet, для работы получает название csv файла, с которым мы работаем
        Args:
             file_name (str): Название файла со статистикой
        """
        self.file_name = file_name
        self.vacancies_objects = DataSet.parser_csv(file_name)

    @staticmethod
    def csv_filer(reader, list_naming):
        """Функция считывающая данные из csv файла статистических данных. Также чистит от html-тегов, пробелов и None'ов
        Attributes:
            reader(list[list[str]]):  Исходные данные из csv файла (вакансии)
            list_naming(list[str]): Названия столбцов
        Returns:
             resultList(list[list[str]]): DataSet со всеми вакансиями
        """
        dataList, headerList = reader, list_naming
        if dataList == [] and headerList == []:
            resultList = []
            return resultList
        for i in range(len(dataList)):
            if '' in dataList[i] or len(dataList[i]) < len(headerList):
                dataList[i].clear()
        resultList = list(filter(None, dataList))
        for i in range(len(resultList)):
            tempList = []
            for j in range(len(resultList[i])):
                if j == 2 and resultList[i][j]:
                    resultList[i][j] = resultList[i][j].split('\n')
                elif j == 2 and not resultList[i][j]:
                    resultList[i][j] = []
                else:
                    resultList[i][j] = re.sub('<.*?>', '', resultList[i][j]).replace("    ", " ").replace("  ",
                                                                                                          " ").replace(
                        "  ", " ").replace(
                        "  ", " ").rstrip().lstrip().replace("  ", " ").replace(" ", " ")
                tempList.append(resultList[i][j])
            resultList[i] = tempList
        return resultList

    @staticmethod
    def reader_csv(file_name):
        """ Функция читает CSV файл статистики и создает два list'a с данными и с заголовками файла csv

        Attributes:
            file_name (str): Имя файла CSV, из которого будут читаться данные
        Return:
            list: Лист с данными вакансий
            list: Лист с заголовками csv файла
        """
        with open(file_name, encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            dataList = list(reader)
            if dataList == []:
                headerList = dataList
                return dataList, headerList
            else:
                headerList = dataList[0]
                dataList.pop(0)
                return dataList, headerList

    @staticmethod
    def get_vacancies_list(result_list, title_list):
        """ Функция получает лист вакансий, состоящий из объектов Vacancy

        Attributes:
            result_list(list): Лист с данными о вакансиях
            title_list(list): Лист с заголовками
        Return:
            list: Лист вакансий с объектами вакансий
        """
        vacancies_list = []
        if result_list:
            for row in result_list:
                dict = {}
                for i in range(len(row)):
                    dict[title_list[i]] = row[i]
                vacancies_list.append(
                    Vacancy(dict['name'], Salary(dict['salary_from'], dict['salary_to'], dict['salary_currency']),
                            dict['area_name'], dict['published_at']))
            return vacancies_list
        else:
            return vacancies_list

    @staticmethod
    def parser_csv(file_name):
        """ Функция обрабатывает файл csv и получает лист вакансий с объектами вакансий

        Attributes:
            file_name(str): Название csv файла со статистикой
        Return:
            list: Лист вакансий с объектами вакансий
        """
        data_list, title_list = DataSet.reader_csv(file_name)
        result_list = DataSet.csv_filer(data_list, title_list)
        vacancies_list = DataSet.get_vacancies_list(result_list, title_list)
        return vacancies_list

def main():
    """ Функция запускает обработку данных и получает данные в нужном виде(pdf-файл)
    """
    InputConect().start()
